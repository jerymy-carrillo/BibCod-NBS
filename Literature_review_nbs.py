#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
BibCod-NBS
Bibliographic Coding Tool for Nature-Based Solutions Systematic Reviews
================================================================================

Version:        1.0
Language:       Python 3.12.3
Dependencies:   openpyxl >= 3.1.0

--------------------------------------------------------------------------------
PURPOSE
--------------------------------------------------------------------------------
This script was developed ad hoc to support a systematic mapping review on
Nature-Based Solutions (NBS) applied to water management, conducted within
the framework of the CYTED Network (Ibero-American Programme of Science and
Technology for Development) and led by the Department of Geology, University
of Oviedo (Spain).

The tool performs three sequential tasks on a bibliographic corpus exported
from Scopus:

    1.  Parses a RIS (.ris) bibliographic file and extracts basic metadata
        (authors, title, year, journal, DOI) for each record.
    2.  Provides the full structured coding protocol (10 thematic blocks,
        43 variables) used by the research team to manually code each
        article after full-text reading.
    3.  Computes, for every article, a thematic affinity score (the "DS
        Index") with the journal collection "Nature-based Solutions applied
        to water management" (Discover Sustainability, Springer), and
        exports a fully formatted Excel workbook containing the coded
        database, the prioritised sub-corpus, descriptive statistics, and
        the coding protocol itself.

--------------------------------------------------------------------------------
METHODOLOGICAL NOTE ON THE CODING PROCESS
--------------------------------------------------------------------------------
The 43 coding variables described in this script (see PROTOCOL_BLOCKS below)
require expert judgement based on full-text reading of each article. This
tool does NOT automate that judgement: it automates the DATA STRUCTURE,
the DS Index CALCULATION, and the REPORTING/EXPORT steps. The qualitative
coding itself was carried out independently by at least two members of the
research team for every one of the 85 articles in the corpus, with
discrepancies resolved by consensus, following good-practice guidance for
data extraction forms in systematic reviews (Buchter et al., 2020) and
semi-automated extraction workflows (Schmidt et al., 2025).

--------------------------------------------------------------------------------
REFERENCES
--------------------------------------------------------------------------------
Buchter, R. B., Weise, A., & Pieper, D. (2020). Development, testing and use
    of data extraction forms in systematic reviews: a review of methodological
    guidance. BMC Medical Research Methodology, 20, 259.
    https://doi.org/10.1186/s12874-020-01143-3

Page, M. J., McKenzie, J. E., Bossuyt, P. M., et al. (2021). The PRISMA 2020
    statement: an updated guideline for reporting systematic reviews. BMJ,
    372, n71. https://doi.org/10.1136/bmj.n71

Schmidt, L., et al. (2025). Data extraction methods for systematic review
    (semi)automation: update of a living systematic review [version 3].
    F1000Research, 10, 401. https://doi.org/10.12688/f1000research.51117.3

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------
    python bibcod_nbs.py --input bibliography.ris --output results.xlsx

Command-line arguments:
    --input,  -i   Path to the RIS file exported from Scopus
                   (default: "bibliography.ris")
    --output, -o   Path to the Excel file to be generated
                   (default: "bibcod_nbs_results.xlsx")

--------------------------------------------------------------------------------
OUTPUT STRUCTURE (Excel workbook)
--------------------------------------------------------------------------------
    Sheet 1  "Full Analysis"        Complete coded database (n articles x 45 cols)
    Sheet 2  "TOP Articles DS>=7"   Prioritised sub-corpus ranked by DS Index
    Sheet 3  "Descriptive Stats"    Frequencies and percentages by thematic block
    Sheet 4  "Coding Protocol"      Full protocol: 10 blocks, 43 variables

================================================================================
"""

# ─────────────────────────────────────────────────────────────────────────────
# IMPORTS
# ─────────────────────────────────────────────────────────────────────────────
import os
import sys
import argparse
from collections import Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: The 'openpyxl' library is not installed.")
    print("Install it with: pip install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 1 — RIS FILE PARSING
# ─────────────────────────────────────────────────────────────────────────────

import re

# Regex pattern for a RIS tag line: two uppercase letters/digits, whitespace,
# a hyphen, whitespace, then the (possibly empty) value.
# Using a regex instead of fixed-width slicing makes the parser robust to
# lines with empty values (e.g. "ER  - " or "VL  - "), where naive string
# slicing after .strip() would otherwise misalign the fixed-width tag format.
_RIS_TAG_PATTERN = re.compile(r"^([A-Z0-9]{2})\s+-\s*(.*)$")


def read_ris_file(file_path):
    """
    Read and parse a bibliographic file in RIS format (.ris).

    RIS is a standardised tagged file format used for exchanging
    bibliographic references between citation databases (e.g. Scopus,
    Web of Science) and reference management software (e.g. Zotero,
    Mendeley, EndNote). Each bibliographic record begins with a document
    type tag (TY) and ends with an end-of-record tag (ER).

    Relevant tags extracted by this parser:
        TY  - Document type   (e.g. JOUR = journal article)
        AU  - Author          (repeatable; one tag per author)
        TI  - Title           (article title)
        PY  - Publication year
        JO  - Journal name
        DO  - DOI (Digital Object Identifier)
        ER  - End of record marker

    Implementation note
    --------------------
    Tag lines are matched with a regular expression rather than fixed-width
    string slicing. This avoids a subtle parsing bug that occurs when a
    field has an empty value (e.g. "ER  - " or "VL  - "): stripping
    trailing whitespace from such a line shortens it, which breaks any
    parser relying on a fixed character offset (e.g. line[2:6]) to locate
    the " - " separator. The regex r"^([A-Z0-9]{2})\\s+-\\s*(.*)$" matches
    the tag format regardless of value length or trailing whitespace.

    Parameters
    ----------
    file_path : str
        Path to the .ris file exported from Scopus (or any RIS-compliant
        source).

    Returns
    -------
    list[dict]
        A list of dictionaries, one per bibliographic record, each with
        the keys: "authors" (list[str]), "title" (str), "year" (int or
        None), "journal" (str), "doi" (str).

    Raises
    ------
    FileNotFoundError
        If the specified file path does not exist.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    records = []
    current_record = {
        "authors": [], "title": "", "year": None,
        "journal": "", "doi": ""
    }

    with open(file_path, "r", encoding="utf-8") as f:
        for raw_line in f:
            line = raw_line.rstrip("\n\r")
            if not line.strip():
                continue

            match = _RIS_TAG_PATTERN.match(line)
            if not match:
                continue

            tag, value = match.group(1), match.group(2).strip()

            if tag == "AU":
                current_record["authors"].append(value)
            elif tag == "TI":
                current_record["title"] = value
            elif tag == "PY":
                try:
                    current_record["year"] = int(value[:4])
                except (ValueError, IndexError):
                    current_record["year"] = None
            elif tag == "JO":
                current_record["journal"] = value
            elif tag == "DO":
                current_record["doi"] = value
            elif tag == "ER":
                # End of record: store it if it has a title, then reset
                if current_record["title"]:
                    records.append(current_record.copy())
                current_record = {
                    "authors": [], "title": "", "year": None,
                    "journal": "", "doi": ""
                }

    print(f"[OK] RIS file parsed: {len(records)} records found")
    return records


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 2 — STRUCTURED CODING PROTOCOL (10 BLOCKS / 43 VARIABLES)
# ─────────────────────────────────────────────────────────────────────────────
#
# The coding protocol is defined as a list of structured questions organised
# into 10 thematic blocks. Each variable record includes:
#
#   block       : Thematic block number and name
#   variable    : Internal variable name (used as dictionary key elsewhere)
#   question    : Operative question presented to the human coder
#   var_type    : Variable type (dichotomous / categorical / ordinal)
#                 "(MR)" denotes multiple-response categorical variables,
#                 i.e. one article may belong to more than one category.
#   values      : Set of admissible values for the variable
#
# This protocol was developed by the research team through an iterative
# process of expert consensus and pilot coding, prior to its systematic
# application to the full corpus (n = 85 articles).
# ─────────────────────────────────────────────────────────────────────────────

PROTOCOL_BLOCKS = [
    # ── BLOCK 1: PUBLICATION METADATA ────────────────────────────────────────
    {
        "block": "1. Publication metadata",
        "variable": "year",
        "question": "In which year was the article published?",
        "var_type": "numerical",
        "values": "Year (integer)"
    },
    {
        "block": "1. Publication metadata",
        "variable": "journal",
        "question": "In which journal was the article published?",
        "var_type": "categorical",
        "values": "Journal name"
    },
    {
        "block": "1. Publication metadata",
        "variable": "quartile",
        "question": "What is the journal quartile according to JCR/Scimago?",
        "var_type": "ordinal",
        "values": "Q1 / Q2 / Q3 / Q4 / Conference / Not indexed / Book"
    },

    # ── BLOCK 2: ARTICLE TYPE AND INTERVENTION ────────────────────────────────
    {
        "block": "2. Article type and intervention",
        "variable": "article_type",
        "question": "Is the article empirical research, a review, conceptual, or a book chapter?",
        "var_type": "categorical",
        "values": "Research / Review / Conceptual / Book chapter"
    },
    {
        "block": "2. Article type and intervention",
        "variable": "intervention",
        "question": "Does it present a real implemented intervention, a design proposal, or a purely theoretical analysis?",
        "var_type": "categorical",
        "values": "Real / Proposal / Theoretical / Real+Proposal"
    },
    {
        "block": "2. Article type and intervention",
        "variable": "case_study",
        "question": "Does the article develop a specific case study?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },

    # ── BLOCK 3: METHODOLOGICAL RIGOUR ────────────────────────────────────────
    {
        "block": "3. Methodological rigour",
        "variable": "clear_methodology",
        "question": "Is the methodology described in sufficient detail to be replicated?",
        "var_type": "ordinal",
        "values": "Yes / Partial / No"
    },
    {
        "block": "3. Methodological rigour",
        "variable": "novel_methodology",
        "question": "Does the article propose a new methodology or apply an established one?",
        "var_type": "dichotomous",
        "values": "New / Existing"
    },
    {
        "block": "3. Methodological rigour",
        "variable": "data_quality",
        "question": "Are the data used complete, valid, and verifiable?",
        "var_type": "ordinal",
        "values": "High / Moderate / Low"
    },
    {
        "block": "3. Methodological rigour",
        "variable": "limitations_stated",
        "question": "Does the article explicitly identify its own limitations?",
        "var_type": "ordinal",
        "values": "Yes / Partial / No"
    },

    # ── BLOCK 4: GEOGRAPHIC AND SOCIAL CONTEXT ────────────────────────────────
    {
        "block": "4. Geographic and social context",
        "variable": "country_region",
        "question": "In which country or region is the study conducted?",
        "var_type": "categorical",
        "values": "Specific country and region"
    },
    {
        "block": "4. Geographic and social context",
        "variable": "global_zone",
        "question": "To which global geographic zone does the case study belong?",
        "var_type": "categorical",
        "values": "Europe / Asia-Pacific / Latin America / North America / Africa / International"
    },
    {
        "block": "4. Geographic and social context",
        "variable": "community_type",
        "question": "Is the setting urban, peri-urban, rural, or coastal?",
        "var_type": "categorical (MR)",
        "values": "Urban / Peri-urban / Rural / Coastal"
    },
    {
        "block": "4. Geographic and social context",
        "variable": "vulnerable_community",
        "question": "Does the study explicitly address socially vulnerable or at-risk communities?",
        "var_type": "ordinal",
        "values": "Yes / Partial / No"
    },
    {
        "block": "4. Geographic and social context",
        "variable": "spatial_scale",
        "question": "At which spatial scale are the NBS applied?",
        "var_type": "categorical",
        "values": "Plot / Neighbourhood / Municipal / Catchment / Regional / National / Multiple"
    },

    # ── BLOCK 5: HYDROLOGY AND GEOLOGICAL DISCIPLINE ──────────────────────────
    {
        "block": "5. Hydrology and geological discipline",
        "variable": "geological_discipline",
        "question": "Does the article incorporate an explicit geological or hydrogeological discipline? Which one?",
        "var_type": "categorical",
        "values": "Hydrology / Hydrogeology / Geomorphology / Coastal geology / GIS-Remote sensing / Aquatic ecology / None"
    },
    {
        "block": "5. Hydrology and geological discipline",
        "variable": "water_system",
        "question": "What type of water system is the object of study?",
        "var_type": "categorical (MR)",
        "values": "Surface runoff / River-catchment / Aquifer / Wetland / Coastal-estuarine / Urban drainage / Integrated cycle"
    },
    {
        "block": "5. Hydrology and geological discipline",
        "variable": "integrated_management",
        "question": "Does the article address water from an integrated resource management perspective?",
        "var_type": "ordinal",
        "values": "Yes / Partial / No"
    },

    # ── BLOCK 6: NBS TYPOLOGY ──────────────────────────────────────────────────
    {
        "block": "6. NBS typology",
        "variable": "suds",
        "question": "Does it apply Sustainable Urban Drainage Systems (green roofs, permeable pavements, bioretention)?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "6. NBS typology",
        "variable": "river_restoration",
        "question": "Does it address river, channel, or riverbank restoration as NBS?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "6. NBS typology",
        "variable": "catchment_restoration",
        "question": "Does it apply NBS at the river-catchment scale?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "6. NBS typology",
        "variable": "wetlands",
        "question": "Does it use wetlands, marshes, or mangroves as NBS?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "6. NBS typology",
        "variable": "reforestation",
        "question": "Does it include reforestation, riparian forests, or urban tree cover as NBS?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "6. NBS typology",
        "variable": "urban_agriculture",
        "question": "Does it address urban or peri-urban agriculture as water-related NBS?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "6. NBS typology",
        "variable": "traditional_practices",
        "question": "Does it incorporate or value historical hydraulic practices or ancestral knowledge?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },

    # ── BLOCK 7: WATER-RELATED PROBLEM ADDRESSED ──────────────────────────────
    {
        "block": "7. Water-related problem addressed",
        "variable": "floods",
        "question": "Does it address water management from a flood-risk perspective?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "7. Water-related problem addressed",
        "variable": "droughts",
        "question": "Does it address water scarcity or drought management?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "7. Water-related problem addressed",
        "variable": "water_quality",
        "question": "Does it address water pollution or water quality deterioration?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "7. Water-related problem addressed",
        "variable": "aquifer_recharge",
        "question": "Does it address aquifer recharge or infiltration systems as NBS?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },

    # ── BLOCK 8: CLIMATE CHANGE ────────────────────────────────────────────────
    {
        "block": "8. Climate change",
        "variable": "climate_change",
        "question": "Does it analyse NBS in relation to climate change or extreme climatic events?",
        "var_type": "ordinal",
        "values": "Yes / Partial / No"
    },
    {
        "block": "8. Climate change",
        "variable": "urban_microclimate",
        "question": "Does it address the effects of NBS on the urban microclimate?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "8. Climate change",
        "variable": "monitoring",
        "question": "Does it include monitoring or time-series follow-up data on the implemented NBS?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },

    # ── BLOCK 9: PLANNING AND REGULATORY FRAMEWORK ────────────────────────────
    {
        "block": "9. Planning and regulatory framework",
        "variable": "spatial_planning",
        "question": "Does it integrate NBS into spatial or urban planning frameworks?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "9. Planning and regulatory framework",
        "variable": "planning_instruments",
        "question": "Does it employ or propose formal planning instruments?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "9. Planning and regulatory framework",
        "variable": "urban_conflicts",
        "question": "Does it address conflicts between NBS and existing urban development models?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "9. Planning and regulatory framework",
        "variable": "policy_regulation",
        "question": "Does it propose or analyse regulatory frameworks or public policies for NBS?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "9. Planning and regulatory framework",
        "variable": "grey_infrastructure_comparison",
        "question": "Does it compare the effectiveness of NBS with conventional grey infrastructure?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },

    # ── BLOCK 10: SOCIAL DIMENSION ─────────────────────────────────────────────
    {
        "block": "10. Social dimension",
        "variable": "citizen_participation",
        "question": "Does it incorporate citizen participation or co-design processes?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "10. Social dimension",
        "variable": "social_acceptance",
        "question": "Does it analyse the social perception or acceptance of the NBS?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "10. Social dimension",
        "variable": "traditional_knowledge",
        "question": "Does it incorporate or valorise local or traditional knowledge?",
        "var_type": "dichotomous",
        "values": "Yes / No"
    },
    {
        "block": "10. Social dimension",
        "variable": "practical_applicability",
        "question": "Do the results have direct practical applicability?",
        "var_type": "ordinal",
        "values": "High / Medium / Low"
    },
    {
        "block": "10. Social dimension",
        "variable": "transferability",
        "question": "Are the results transferable to other geographic contexts?",
        "var_type": "ordinal",
        "values": "High / Medium / Low"
    },
]


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 3 — DS INDEX CALCULATION
# ─────────────────────────────────────────────────────────────────────────────

def compute_ds_index(article):
    """
    Compute the thematic affinity score ("DS Index") for a single article.

    The DS Index is a composite score derived from weighted expert
    assessment, designed to measure how closely an article aligns with the
    thematic scope of the "Nature-based Solutions applied to water
    management" collection of the journal Discover Sustainability
    (Springer, Impact Factor ~3.0).

    The score is built from six weighted criteria, summing to a maximum of
    10 points:

        Criterion 1 (weight 3): NBS + water management as the central focus
            Checks whether the article treats NBS as a primary subject AND
            water management as a central problem (not peripheral).

        Criterion 2 (weight 2): Integrated water resource management approach
            Checks whether water is addressed from a systemic perspective
            rather than a single isolated issue (e.g. floods only).

        Criterion 3 (weight 2): Explicit sustainability/resilience framework
            Inferred from the combined presence of spatial planning,
            climate change consideration, and high practical applicability.

        Criterion 4 (weight 1): Climate change as driver or context
            Checks whether climate change is explicitly mentioned as a
            motivation or contextual frame for the study.

        Criterion 5 (weight 1): Social dimension, governance or vulnerability
            Checks for the presence of citizen participation, vulnerable
            communities, or governance analysis.

        Criterion 6 (weight 1): Practical applicability and transferability
            Checks whether the results are reported as both highly
            applicable and highly transferable.

    Interpretation scale:
        9-10  : Optimal alignment with the DS collection scope
        7-8   : High alignment
        5-6   : Moderate alignment
        1-4   : Low or peripheral alignment

    Parameters
    ----------
    article : dict
        Dictionary containing the coded variables for a single article,
        using the variable names defined in PROTOCOL_BLOCKS.

    Returns
    -------
    int
        DS Index score on a 1-10 scale.
    """
    score = 0

    # ── Criterion 1: NBS + water as central focus (weight 3) ─────────────────
    has_nbs = any([
        article.get("suds") == "Yes",
        article.get("river_restoration") == "Yes",
        article.get("catchment_restoration") == "Yes",
        article.get("wetlands") == "Yes",
        article.get("reforestation") == "Yes",
    ])
    has_water_issue = any([
        article.get("floods") == "Yes",
        article.get("droughts") == "Yes",
        article.get("water_quality") == "Yes",
        article.get("integrated_management") == "Yes",
        article.get("aquifer_recharge") == "Yes",
    ])
    if has_nbs and has_water_issue:
        score += 3
    elif has_nbs or has_water_issue:
        score += 1

    # ── Criterion 2: Integrated resource management (weight 2) ───────────────
    if article.get("integrated_management") == "Yes":
        score += 2
    elif article.get("integrated_management") == "Partial":
        score += 1

    # ── Criterion 3: Sustainability / resilience framework (weight 2) ────────
    resilience_indicators = sum([
        1 if article.get("spatial_planning") == "Yes" else 0,
        1 if article.get("climate_change") in ["Yes", "Partial"] else 0,
        1 if article.get("practical_applicability") == "High" else 0,
    ])
    if resilience_indicators >= 2:
        score += 2
    elif resilience_indicators == 1:
        score += 1

    # ── Criterion 4: Climate change as driver (weight 1) ──────────────────────
    if article.get("climate_change") == "Yes":
        score += 1

    # ── Criterion 5: Social dimension / governance (weight 1) ────────────────
    has_social_dimension = any([
        article.get("citizen_participation") == "Yes",
        article.get("social_acceptance") == "Yes",
        article.get("vulnerable_community") == "Yes",
        article.get("traditional_knowledge") == "Yes",
    ])
    if has_social_dimension:
        score += 1

    # ── Criterion 6: Applicability and transferability (weight 1) ────────────
    if (article.get("practical_applicability") == "High" and
            article.get("transferability") == "High"):
        score += 1

    # ── Normalise to the 1-10 scale ───────────────────────────────────────────
    # Maximum attainable raw score: 3+2+2+1+1+1 = 10
    # The lower bound is fixed at 1 (no article scores zero)
    ds_index = max(1, min(10, score))
    return ds_index


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 4 — EXCEL WORKBOOK GENERATION
# ─────────────────────────────────────────────────────────────────────────────

def _build_style_kit():
    """
    Build and return the dictionary of visual styles used across the
    generated Excel workbook.

    Colour coding follows a consistent semantic logic throughout the
    workbook:
        Dark blue   (#1F4E79) - Main titles and primary headers
        Medium blue (#2E75B6) - Metadata and secondary headers
        Green       (#70AD47) - Positive indicators / high quality
        Amber       (#FFC000) - Medium / moderate indicators
        Red         (#C00000) - Critical gaps / low values
        Purple      (#7030A0) - DS Index and editorial framing

    Returns
    -------
    dict
        A dictionary exposing reusable Font, Alignment, Border and fill
        helper objects for cell formatting.
    """
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    thin_side = Side(style="thin", color="CCCCCC")
    border = Border(left=thin_side, right=thin_side,
                     top=thin_side, bottom=thin_side)

    return {
        "fill": fill,
        "border": border,
        "center": Alignment(horizontal="center", vertical="center",
                             wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center",
                           wrap_text=True),
        "font_title": Font(name="Arial", bold=True, color="FFFFFF", size=10),
        "font_header": Font(name="Arial", bold=True, color="FFFFFF", size=9),
        "font_body": Font(name="Arial", size=8),
        "font_bold": Font(name="Arial", bold=True, size=8),
    }


def generate_excel_report(articles, output_path="bibcod_nbs_results.xlsx"):
    """
    Generate the complete four-sheet Excel report from the coded corpus.

    The resulting database has a matrix structure:
        n_articles rows x n_variables columns
    (85 x 43 in the original review corpus, plus the computed DS Index)

    Parameters
    ----------
    articles : list[dict]
        List of article dictionaries, each containing the metadata fields
        extracted from RIS plus the 43 coded variables and the computed
        "ds_index" field.
    output_path : str
        File path for the Excel workbook to be created.

    Output sheets
    -------------
    Sheet 1 "Full Analysis"
        The complete coded database with all 43 protocol variables plus
        metadata and the DS Index, colour-coded for quick visual scanning.

    Sheet 2 "TOP Articles DS>=7"
        The prioritised sub-corpus: articles with DS Index >= 7, ranked in
        descending order, intended to support the qualitative synthesis in
        the Results and Discussion sections.

    Sheet 3 "Descriptive Stats"
        Frequency counts and percentages for the key variables, organised
        by thematic block, including flags for identified research gaps.

    Sheet 4 "Coding Protocol"
        The full structured coding protocol (10 blocks, 43 variables) as
        defined in PROTOCOL_BLOCKS, suitable for inclusion as a
        supplementary appendix.
    """
    style = _build_style_kit()
    wb = Workbook()

    # ── SHEET 1: Full Analysis ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Full Analysis"

    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value = (
        f"BibCod-NBS - Coded database | n = {len(articles)} articles | "
        f"10 blocks | 43 variables"
    )
    title_cell.font = style["font_title"]
    title_cell.fill = style["fill"]("1F4E79")
    title_cell.alignment = style["center"]
    ws1.row_dimensions[1].height = 28

    columns = [
        ("ID", 5), ("Authors", 20), ("Title", 45), ("Year", 6),
        ("Journal", 22), ("Quartile", 8), ("Country/Region", 20),
        ("Global Zone", 18), ("Article Type", 16),
        ("Intervention", 14), ("Clear Methodology", 12),
        ("Novel Methodology", 12), ("Data Quality", 10),
        ("Limitations Stated", 12), ("Community Type", 14),
        ("Vulnerable", 10), ("Spatial Scale", 14),
        ("Geological Discipline", 20), ("Water System", 20),
        ("Integrated Mgmt.", 12), ("SUDS", 6),
        ("River Restoration", 12), ("Catchment Restoration", 14),
        ("Wetlands", 8), ("Reforestation", 12),
        ("Urban Agriculture", 12), ("Traditional Practices", 14),
        ("Floods", 8), ("Droughts", 8),
        ("Water Quality", 10), ("Aquifer Recharge", 12),
        ("Climate Change", 12), ("Microclimate", 10),
        ("Monitoring", 10), ("Spatial Planning", 12),
        ("Planning Instr.", 10), ("Urban Conflicts", 10),
        ("Policy/Regulation", 12), ("Grey Infr. Comparison", 14),
        ("Citizen Participation", 14), ("Social Acceptance", 12),
        ("Traditional Knowledge", 14), ("Applicability", 10),
        ("Transferability", 12), ("DS Index (1-10)", 12),
    ]

    header_colours = {
        "ID": "1F4E79",
        "Authors": "2E75B6", "Title": "2E75B6", "Year": "2E75B6",
        "Journal": "2E75B6", "Quartile": "2E75B6",
        "Country/Region": "ED7D31", "Global Zone": "ED7D31",
        "Article Type": "548235", "Intervention": "548235",
        "Clear Methodology": "70AD47", "Novel Methodology": "70AD47",
        "Data Quality": "70AD47", "Limitations Stated": "70AD47",
        "Community Type": "ED7D31", "Vulnerable": "ED7D31",
        "Spatial Scale": "ED7D31",
        "Geological Discipline": "C00000", "Water System": "C00000",
        "Integrated Mgmt.": "C00000",
    }

    for col_idx, (name, width) in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        cell = ws1.cell(2, col_idx, name)
        colour = header_colours.get(name, "4472C4")
        cell.font = style["font_header"]
        cell.fill = style["fill"](colour)
        cell.alignment = style["center"]
        cell.border = style["border"]
        ws1.column_dimensions[col_letter].width = width
    ws1.row_dimensions[2].height = 36

    field_order = [
        "id", "authors", "title", "year", "journal", "quartile",
        "country_region", "global_zone", "article_type", "intervention",
        "clear_methodology", "novel_methodology", "data_quality",
        "limitations_stated", "community_type", "vulnerable_community",
        "spatial_scale", "geological_discipline", "water_system",
        "integrated_management", "suds", "river_restoration",
        "catchment_restoration", "wetlands", "reforestation",
        "urban_agriculture", "traditional_practices", "floods", "droughts",
        "water_quality", "aquifer_recharge", "climate_change",
        "urban_microclimate", "monitoring", "spatial_planning",
        "planning_instruments", "urban_conflicts", "policy_regulation",
        "grey_infrastructure_comparison", "citizen_participation",
        "social_acceptance", "traditional_knowledge",
        "practical_applicability", "transferability", "ds_index",
    ]

    for i, article in enumerate(articles):
        row_idx = i + 3
        is_even = i % 2 == 0
        row_bg = "EBF3E8" if is_even else "FFFFFF"

        for col_idx, field in enumerate(field_order, 1):
            value = article.get(field, "")
            cell = ws1.cell(row_idx, col_idx, value)
            cell.font = style["font_body"]
            cell.alignment = style["left"]
            cell.border = style["border"]
            cell.fill = style["fill"](row_bg)

            # Highlight the DS Index column with a quality colour scale
            if field == "ds_index" and isinstance(value, (int, float)):
                if value >= 8:
                    cell.fill = style["fill"]("70AD47")
                    cell.font = Font(name="Arial", size=8, bold=True,
                                      color="FFFFFF")
                elif value >= 6:
                    cell.fill = style["fill"]("FFC000")
                elif value < 5:
                    cell.fill = style["fill"]("FF9999")

        ws1.row_dimensions[row_idx].height = 28

    ws1.freeze_panes = "C3"

    # ── SHEET 2: TOP Articles (DS Index >= 7) ─────────────────────────────────
    ws2 = wb.create_sheet("TOP Articles DS>=7")
    top_articles = sorted(
        [a for a in articles if a.get("ds_index", 0) >= 7],
        key=lambda a: -a.get("ds_index", 0)
    )

    ws2.merge_cells("A1:F1")
    cell = ws2["A1"]
    cell.value = (
        f"Prioritised sub-corpus: articles with DS Index >= 7 | "
        f"n = {len(top_articles)} articles"
    )
    cell.font = style["font_title"]
    cell.fill = style["fill"]("7030A0")
    cell.alignment = style["center"]
    ws2.row_dimensions[1].height = 24

    headers_2 = ["Rank", "Authors", "Title", "Year", "Journal",
                 "DS Index", "NBS Type", "Water Problem",
                 "Global Zone", "Vulnerable", "Climate Change"]
    for ci, h in enumerate(headers_2, 1):
        c = ws2.cell(2, ci, h)
        c.font = style["font_header"]
        c.fill = style["fill"]("7030A0")
        c.alignment = style["center"]
        c.border = style["border"]

    nbs_labels = {
        "suds": "SUDS",
        "river_restoration": "River restoration",
        "catchment_restoration": "Catchment restoration",
        "wetlands": "Wetlands",
        "reforestation": "Reforestation",
        "traditional_practices": "Traditional practices",
    }

    for rank, article in enumerate(top_articles, 1):
        row_idx = rank + 2
        nbs_present = [
            label for field, label in nbs_labels.items()
            if article.get(field) == "Yes"
        ]

        row_data = [
            rank,
            article.get("authors", ""),
            article.get("title", ""),
            article.get("year", ""),
            article.get("journal", ""),
            article.get("ds_index", ""),
            ", ".join(nbs_present) if nbs_present else "Not specified",
            article.get("water_problem", ""),
            article.get("global_zone", ""),
            article.get("vulnerable_community", ""),
            article.get("climate_change", ""),
        ]
        for ci, value in enumerate(row_data, 1):
            c = ws2.cell(row_idx, ci, value)
            c.font = style["font_body"]
            c.alignment = style["left"]
            c.border = style["border"]
            if ci == 6 and isinstance(value, int) and value >= 9:
                c.fill = style["fill"]("70AD47")
                c.font = Font(name="Arial", size=8, bold=True,
                               color="FFFFFF")
        ws2.row_dimensions[row_idx].height = 24

    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 48
    ws2.column_dimensions["D"].width = 6
    ws2.column_dimensions["E"].width = 24
    ws2.freeze_panes = "C3"

    # ── SHEET 3: Descriptive Statistics ───────────────────────────────────────
    ws3 = wb.create_sheet("Descriptive Stats")
    n_total = len(articles)

    def count_value(field, value):
        return sum(1 for a in articles if a.get(field) == value)

    stats_rows = [
        ["STATISTIC", "n", "%", "Interpretation"],
        ["TOTAL CORPUS", n_total, "100%", "Full analysis baseline"],
        [],
        ["-- TEMPORAL DISTRIBUTION --------------------", "", "", ""],
        ["2017-2020",
         sum(1 for a in articles if (a.get("year") or 0) <= 2020),
         "", "Early-stage field"],
        ["2021-2022",
         sum(1 for a in articles if 2021 <= (a.get("year") or 0) <= 2022),
         "", "Consolidation"],
        ["2023-2024",
         sum(1 for a in articles if 2023 <= (a.get("year") or 0) <= 2024),
         "", "Acceleration"],
        ["2025-2026",
         sum(1 for a in articles if (a.get("year") or 0) >= 2025),
         "", "Current expansion"],
        [],
        ["-- JOURNAL QUARTILES ------------------------", "", "", ""],
        ["Q1", count_value("quartile", "Q1"), "", "High visibility"],
        ["Q2", count_value("quartile", "Q2"), "", "Good visibility"],
        ["Q3", count_value("quartile", "Q3"), "", "Medium visibility"],
        ["Q4", count_value("quartile", "Q4"), "", "Limited visibility"],
        ["Q1+Q2 combined",
         count_value("quartile", "Q1") + count_value("quartile", "Q2"),
         "", "High editorial quality"],
        [],
        ["-- NBS TYPES (multiple response) ------------", "", "", ""],
        ["SUDS", count_value("suds", "Yes"), "", "Urban green infrastructure"],
        ["Wetlands", count_value("wetlands", "Yes"),
         "", "Natural/constructed wetlands"],
        ["River restoration", count_value("river_restoration", "Yes"),
         "", "Channels and riverbanks"],
        ["Reforestation", count_value("reforestation", "Yes"),
         "", "Forests and tree cover"],
        ["Catchment restoration", count_value("catchment_restoration", "Yes"),
         "", "Catchment-scale interventions"],
        ["Traditional practices", count_value("traditional_practices", "Yes"),
         "", "GAP - Ancestral knowledge"],
        [],
        ["-- GEOLOGICAL DIMENSION (critical gap) -------", "", "", ""],
        ["No explicit geological discipline",
         count_value("geological_discipline", "None"),
         "", "GAP - Team's core contribution"],
        ["Hydrogeology",
         count_value("geological_discipline", "Hydrogeology"),
         "", "GAP - Vacant niche"],
        [],
        ["-- SOCIAL DIMENSION --------------------------", "", "", ""],
        ["Vulnerable communities",
         count_value("vulnerable_community", "Yes"),
         "", "GAP - Global South"],
        ["Citizen participation",
         count_value("citizen_participation", "Yes"),
         "", "GAP - Co-design"],
        ["Traditional knowledge",
         count_value("traditional_knowledge", "Yes"),
         "", "GAP - Local knowledge"],
        [],
        ["-- DS INDEX -----------------------------------", "", "", ""],
        ["DS 9-10 (optimal)",
         sum(1 for a in articles if a.get("ds_index", 0) >= 9),
         "", "Maximum alignment with DS"],
        ["DS 7-8 (high)",
         sum(1 for a in articles if 7 <= a.get("ds_index", 0) <= 8),
         "", "High alignment"],
        ["DS 5-6 (moderate)",
         sum(1 for a in articles if 5 <= a.get("ds_index", 0) <= 6),
         "", "Moderate alignment"],
        ["DS 1-4 (low)",
         sum(1 for a in articles if a.get("ds_index", 0) <= 4),
         "", "Low alignment"],
    ]

    for ri, row_data in enumerate(stats_rows, 1):
        if not row_data:
            continue
        for ci, value in enumerate(row_data, 1):
            cell = ws3.cell(ri, ci, value)
            if ri == 1:
                cell.font = style["font_header"]
                cell.fill = style["fill"]("1F4E79")
            elif str(value).startswith("--"):
                cell.font = Font(name="Arial", bold=True, size=9)
                cell.fill = style["fill"]("D9E1F2")
            else:
                cell.font = style["font_body"]
                if ri % 2 == 0:
                    cell.fill = style["fill"]("F2F2F2")
            cell.alignment = style["left"]
            cell.border = style["border"]

        # Compute percentage for data rows with a numeric count
        if (len(row_data) >= 2 and isinstance(row_data[1], int)
                and row_data[1] >= 0 and ri > 1
                and not str(row_data[0]).startswith("--")
                and n_total > 0):
            pct = round(row_data[1] / n_total * 100, 1)
            ws3.cell(ri, 3, f"{pct}%")

    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 8
    ws3.column_dimensions["C"].width = 8
    ws3.column_dimensions["D"].width = 40

    # ── SHEET 4: Coding Protocol ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Coding Protocol")

    ws4.merge_cells("A1:E1")
    cell = ws4["A1"]
    cell.value = ("BibCod-NBS - Data extraction and coding protocol | "
                  "10 thematic blocks | 43 variables")
    cell.font = style["font_title"]
    cell.fill = style["fill"]("1F4E79")
    cell.alignment = style["center"]
    ws4.row_dimensions[1].height = 24

    headers_4 = ["Block", "Variable", "Operative Question",
                 "Variable Type", "Possible Values"]
    for ci, h in enumerate(headers_4, 1):
        c = ws4.cell(2, ci, h)
        c.font = style["font_header"]
        c.fill = style["fill"]("2E75B6")
        c.alignment = style["center"]
        c.border = style["border"]

    for ri, var in enumerate(PROTOCOL_BLOCKS, 3):
        for ci, field in enumerate(
            ["block", "variable", "question", "var_type", "values"], 1
        ):
            c = ws4.cell(ri, ci, var[field])
            c.font = style["font_body"]
            c.alignment = style["left"]
            c.border = style["border"]
            if ri % 2 == 0:
                c.fill = style["fill"]("F2F8FF")
        ws4.row_dimensions[ri].height = 30

    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 24
    ws4.column_dimensions["C"].width = 60
    ws4.column_dimensions["D"].width = 18
    ws4.column_dimensions["E"].width = 50
    ws4.freeze_panes = "A3"

    wb.save(output_path)
    print(f"[OK] Excel report generated: {output_path}")
    print(f"     Sheet 1: {len(articles)} articles x 44 columns")
    print(f"     Sheet 2: {len(top_articles)} articles with DS Index >= 7")
    print(f"     Sheet 3: Descriptive statistics")
    print(f"     Sheet 4: Coding protocol (43 variables)")


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 5 — DESCRIPTIVE STATISTICAL ANALYSIS (CONSOLE REPORT)
# ─────────────────────────────────────────────────────────────────────────────

def print_statistical_summary(articles):
    """
    Compute and print a descriptive statistical summary of the coded
    corpus to the console.

    The summary reports absolute and relative frequencies (percentages)
    for a selection of key variables across the coding protocol, sorted
    from the most to the least frequent category within each variable.

    Parameters
    ----------
    articles : list[dict]
        List of coded article dictionaries, each including the computed
        "ds_index" field.
    """
    n_total = len(articles)
    print("\n" + "=" * 70)
    print("DESCRIPTIVE STATISTICAL SUMMARY - BibCod-NBS")
    print(f"Corpus size: n = {n_total} articles")
    print("=" * 70)

    summary_variables = [
        ("Temporal distribution",
         Counter(str(a.get("year", "")) for a in articles)),
        ("Journal quartile",
         Counter(a.get("quartile", "N/A") for a in articles)),
        ("Article type",
         Counter(a.get("article_type", "") for a in articles)),
        ("Global geographic zone",
         Counter(a.get("global_zone", "") for a in articles)),
        ("Climate change",
         Counter(a.get("climate_change", "") for a in articles)),
        ("Vulnerable communities",
         Counter(a.get("vulnerable_community", "") for a in articles)),
        ("Geological discipline (presence)",
         Counter("No" if a.get("geological_discipline") == "None"
                 else "Yes" for a in articles)),
        ("Citizen participation",
         Counter(a.get("citizen_participation", "") for a in articles)),
        ("DS Index distribution",
         Counter(str(a.get("ds_index", "")) for a in articles)),
    ]

    for title, counter in summary_variables:
        print(f"\n-- {title} --")
        for value, count in sorted(counter.items(), key=lambda x: -x[1]):
            pct = round(count / n_total * 100, 1) if n_total else 0
            bar = "#" * int(pct / 3)
            print(f"  {value:<32} {count:>3} ({pct:>5}%)  {bar}")

    # Basic descriptive statistics for the DS Index
    ds_values = [
        a.get("ds_index", 0) for a in articles
        if isinstance(a.get("ds_index"), (int, float))
    ]
    if ds_values:
        mean_ds = round(sum(ds_values) / len(ds_values), 2)
        sorted_ds = sorted(ds_values)
        median_ds = sorted_ds[len(sorted_ds) // 2]
        n_high = sum(1 for v in ds_values if v >= 7)
        pct_high = round(n_high / n_total * 100, 1) if n_total else 0

        print(f"\n-- DS Index - Summary statistics --")
        print(f"  Mean:    {mean_ds}")
        print(f"  Median:  {median_ds}")
        print(f"  Minimum: {min(ds_values)}")
        print(f"  Maximum: {max(ds_values)}")
        print(f"  DS >= 7: {n_high} ({pct_high}%)")

    print("\n" + "=" * 70)


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 6 — MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    """
    Main execution function for the BibCod-NBS script.

    Execution flow
    --------------
    1. Read the RIS file and extract bibliographic metadata for each record.
    2. Initialise the data structure for each article with empty fields
       corresponding to the 43 protocol variables (see PROTOCOL_BLOCKS).
    3. Compute the DS Index for each article (note: this requires the
       43 variables to have been coded beforehand by the research team;
       see the methodological note in the module docstring above).
    4. Generate the four-sheet Excel report.
    5. Print the descriptive statistical summary to the console.

    Important note on the coding step
    ----------------------------------
    Coding the 43 variables requires full-text reading of each article by
    the research team; this is an expert qualitative task that this script
    does not automate. The script provides: (a) the data structure and
    operative questions for that coding task (PROTOCOL_BLOCKS), (b) the DS
    Index calculation logic once coding is complete, and (c) the Excel/
    statistical reporting pipeline.

    In the original review (n = 85 articles), coding was carried out
    independently by at least two members of the research team, with
    discrepancies resolved by consensus, following Buchter et al. (2020)
    and Schmidt et al. (2025).
    """
    parser = argparse.ArgumentParser(
        description="BibCod-NBS: Bibliographic coding tool for NBS systematic reviews"
    )
    parser.add_argument(
        "--input", "-i",
        default="bibliography.ris",
        help="Path to the RIS file exported from Scopus "
             "(default: bibliography.ris)"
    )
    parser.add_argument(
        "--output", "-o",
        default="bibcod_nbs_results.xlsx",
        help="Name of the Excel output file "
             "(default: bibcod_nbs_results.xlsx)"
    )
    args = parser.parse_args()

    print("\n" + "=" * 70)
    print("BibCod-NBS v1.0 - Bibliographic Coding Tool")
    print("Nature-Based Solutions + Water Management Systematic Review")
    print("CYTED Network / University of Oviedo - 2026")
    print("=" * 70 + "\n")

    # ── Step 1: Read RIS file ─────────────────────────────────────────────────
    print(f"[1/4] Reading RIS file: {args.input}")
    try:
        ris_records = read_ris_file(args.input)
    except FileNotFoundError as e:
        print(f"\nERROR: {e}")
        print("Please provide the correct path to the .ris file.")
        print("Usage: python bibcod_nbs.py --input your_file.ris")
        sys.exit(1)

    # ── Step 2: Initialise coding structure from RIS metadata ────────────────
    # NOTE: basic metadata (authors, title, year, journal, DOI) is extracted
    # automatically from the RIS file. The 43 protocol variables must be
    # completed manually by the research team following the operative
    # questions in the Anex/Appendix (see PROTOCOL_BLOCKS, or Sheet 4 of
    # the generated Excel workbook).
    print(f"[2/4] Processing {len(ris_records)} bibliographic records...")

    processed_articles = []
    for i, record in enumerate(ris_records):
        article = {
            "id": i + 1,
            "authors": (
                ", ".join(record["authors"][:3]) +
                (" et al." if len(record["authors"]) > 3 else "")
            ),
            "title": record["title"],
            "year": record["year"],
            "journal": record["journal"],
            "doi": record["doi"],
            # The following fields correspond to the 43 protocol variables
            # and must be filled in manually by the coding team:
            "quartile": "",                       # Block 1
            "article_type": "",                   # Block 2
            "intervention": "",                    # Block 2
            "clear_methodology": "",               # Block 3
            "novel_methodology": "",               # Block 3
            "data_quality": "",                    # Block 3
            "limitations_stated": "",              # Block 3
            "country_region": "",                  # Block 4
            "global_zone": "",                     # Block 4
            "community_type": "",                  # Block 4
            "vulnerable_community": "",             # Block 4
            "spatial_scale": "",                   # Block 4
            "geological_discipline": "",           # Block 5
            "water_system": "",                    # Block 5
            "integrated_management": "",           # Block 5
            "suds": "",                             # Block 6
            "river_restoration": "",                # Block 6
            "catchment_restoration": "",            # Block 6
            "wetlands": "",                         # Block 6
            "reforestation": "",                    # Block 6
            "urban_agriculture": "",                # Block 6
            "traditional_practices": "",            # Block 6
            "floods": "",                           # Block 7
            "droughts": "",                         # Block 7
            "water_quality": "",                    # Block 7
            "aquifer_recharge": "",                 # Block 7
            "climate_change": "",                   # Block 8
            "urban_microclimate": "",               # Block 8
            "monitoring": "",                       # Block 8
            "spatial_planning": "",                 # Block 9
            "planning_instruments": "",             # Block 9
            "urban_conflicts": "",                  # Block 9
            "policy_regulation": "",                # Block 9
            "grey_infrastructure_comparison": "",   # Block 9
            "citizen_participation": "",            # Block 10
            "social_acceptance": "",                # Block 10
            "traditional_knowledge": "",            # Block 10
            "practical_applicability": "",          # Block 10
            "transferability": "",                  # Block 10
            "water_problem": "",                    # Summary field
        }
        processed_articles.append(article)

    # ── Step 3: Compute the DS Index ──────────────────────────────────────────
    print("[3/4] Computing DS Index for each article...")
    for article in processed_articles:
        article["ds_index"] = compute_ds_index(article)

    # ── Step 4: Generate Excel report ─────────────────────────────────────────
    print(f"[4/4] Generating Excel report: {args.output}")
    generate_excel_report(processed_articles, args.output)

    # ── Console statistical summary ───────────────────────────────────────────
    print_statistical_summary(processed_articles)

    print(f"\n[DONE] Process completed.")
    print(f"       Output file: {args.output}")
    print(f"       Records processed: {len(processed_articles)}")
    print(f"       Variables per record: 43 (+ DS Index)")
    print(f"\n       NEXT STEP: Complete the manual coding of the 43")
    print(f"       variables in Sheet 1 of the Excel file, following the")
    print(f"       operative questions in the protocol (Sheet 4 / Appendix).")


# ─────────────────────────────────────────────────────────────────────────────
# EXECUTION ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    main()
